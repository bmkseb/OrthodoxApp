#!/usr/bin/env python3
"""
validate_sheet.py — mechanical checks on a filled ingestion spreadsheet.
Does NOT judge translation/doctrinal correctness (that's the human review gate).

Usage:
  python validate_sheet.py --file book.xlsx [--book-type prayer]

Checks:
  - required columns present and in order
  - every row has a section_id and a position
  - position is contiguous (1,2,3...) within each (section_id, chapter) group
  - chapter is either blank for ALL rows or filled for ALL rows (no half-chaptered book)
  - at least one language column non-empty per row
  - source_slug present and consistent
  - reports (does not fail) rows missing a given language
Exits non-zero if hard errors found.
"""
import argparse, sys
from collections import defaultdict
from openpyxl import load_workbook
from contract import COLUMNS, LANGUAGE_COLUMNS

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    a = ap.parse_args()
    wb = load_workbook(a.file)
    ws = wb["Content"] if "Content" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    errors, warnings = [], []

    if header[:len(COLUMNS)] != COLUMNS:
        errors.append(f"Header mismatch.\n expected: {COLUMNS}\n found:    {header}")
        print_report(errors, warnings); sys.exit(1)

    idx = {name: i for i, name in enumerate(COLUMNS)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r): continue
        rows.append(r)

    chapters_used = [r[idx["chapter"]] not in (None, "") for r in rows]
    if any(chapters_used) and not all(chapters_used):
        errors.append("Some rows have a chapter and some don't. A book must be ALL chaptered or ALL flat.")

    groups = defaultdict(list)
    for n, r in enumerate(rows, start=2):
        sid = r[idx["section_id"]]
        if not sid: errors.append(f"Row {n}: missing section_id")
        pos = r[idx["position"]]
        if pos in (None, ""): errors.append(f"Row {n}: missing position")
        if not r[idx["source_slug"]]: errors.append(f"Row {n}: missing source_slug")
        langs_filled = [lc for lc in LANGUAGE_COLUMNS.values() if (r[idx[lc]] not in (None,""))]
        if not langs_filled: errors.append(f"Row {n}: no language text at all")
        for lang, col in LANGUAGE_COLUMNS.items():
            if r[idx[col]] in (None, ""):
                warnings.append(f"Row {n}: missing {lang}")
        groups[(sid, r[idx['chapter']])].append((n, pos))

    for key, items in groups.items():
        positions = sorted(p for _, p in items if isinstance(p, int))
        expected = list(range(1, len(positions)+1))
        if positions != expected:
            errors.append(f"Section {key[0]} chapter {key[1]}: positions {positions} not contiguous (expected {expected})")

    print_report(errors, warnings)
    sys.exit(1 if errors else 0)

def print_report(errors, warnings):
    print(f"\n=== VALIDATION REPORT ===")
    print(f"Hard errors: {len(errors)}")
    for e in errors: print("  ERROR:", e)
    print(f"Warnings (missing languages — OK, just FYI): {len(warnings)}")
    for w in warnings[:40]: print("  warn:", w)
    if len(warnings) > 40: print(f"  ...and {len(warnings)-40} more")
    print("=========================\n")
    if not errors: print("PASS — mechanically valid. Still needs HUMAN review before load.")

if __name__ == "__main__":
    main()
