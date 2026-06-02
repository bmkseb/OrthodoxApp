#!/usr/bin/env python3
"""
generate_template.py — produce a pre-filled ingestion spreadsheet for a book.

Usage:
  python generate_template.py --book-type prayer --book-id <uuid> --out book.xlsx
  python generate_template.py --book-type prayer --book-slug daily-prayer --out book.xlsx

Connects to Supabase (reads SUPABASE_DB_URL env var), looks up the book's
sections, and writes a spreadsheet with locked ID columns pre-filled so a human
only types the content + source_slug. The `chapter` column is included but left
blank — fill it only for chaptered books.
"""
import argparse, os, sys
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from contract import COLUMNS, LOCKED_COLUMNS

def fetch_sections(conn, book_type, book_id=None, book_slug=None):
    cur = conn.cursor()
    if book_type == "prayer":
        if book_slug and not book_id:
            cur.execute("SELECT id FROM prayer_books WHERE slug=%s", (book_slug,))
            row = cur.fetchone()
            if not row: sys.exit(f"No prayer_book with slug '{book_slug}'")
            book_id = row[0]
        cur.execute("""SELECT s.id, s.section_number, s.title_english
                       FROM prayer_sections s WHERE s.book_id=%s
                       ORDER BY s.sort_order, s.section_number""", (book_id,))
        return [(r[0], f"{r[1]}. {r[2]}") for r in cur.fetchall()]
    # Extend here for 'bible' / 'doctrine' as needed.
    sys.exit(f"book-type '{book_type}' not yet wired in generator")

def build(sections, out_path):
    wb = Workbook(); ws = wb.active; ws.title = "Content"
    ws.append(COLUMNS)
    hdr_fill = PatternFill("solid", start_color="1A1511")
    hdr_font = Font(bold=True, color="FFFFFF")
    lock_fill = PatternFill("solid", start_color="E8E4DD")
    thin = Side(style="thin", color="CCCCCC"); border = Border(thin,thin,thin,thin)
    for c in range(1, len(COLUMNS)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = border
    r = 2
    for sid, name in sections:
        ws.cell(row=r, column=1, value=str(sid)).fill = lock_fill
        ws.cell(row=r, column=2, value=name).fill = lock_fill
        ws.cell(row=r, column=4, value=1)  # position default
        for c in range(1, len(COLUMNS)+1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    widths = [38,26,9,9,45,35,35,30,26,30]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"Wrote {out_path} with {len(sections)} section rows.")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--book-type", required=True, choices=["prayer","bible","doctrine"])
    p.add_argument("--book-id"); p.add_argument("--book-slug")
    p.add_argument("--out", required=True)
    a = p.parse_args()
    url = os.environ.get("SUPABASE_DB_URL")
    if not url: sys.exit("Set SUPABASE_DB_URL env var (Supabase connection string).")
    conn = psycopg2.connect(url)
    secs = fetch_sections(conn, a.book_type, a.book_id, a.book_slug)
    build(secs, a.out)
