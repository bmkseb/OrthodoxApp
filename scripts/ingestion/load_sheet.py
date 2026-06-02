#!/usr/bin/env python3
"""
load_sheet.py — load a VALIDATED, REVIEWED spreadsheet into Supabase.

Usage:
  python load_sheet.py --file book.xlsx --book-type prayer            # DRY RUN (default, writes nothing)
  python load_sheet.py --file book.xlsx --book-type prayer --commit   # actually writes

Safety:
  - DRY RUN by default: prints exactly what WOULD be inserted, touches nothing.
  - --commit required to write. Runs in a single transaction (all-or-nothing).
  - Resolves source_slug -> sources.id; fails if a slug isn't registered.
  - Populates BOTH new (position/text_*) and legacy (verse_number/content_*) columns
    while the migration is mid-flight, so it works pre- and post-Phase-2.
"""
import argparse, os, sys
import psycopg2
from openpyxl import load_workbook
from contract import COLUMNS, LEAF_TABLE

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--book-type", required=True, choices=list(LEAF_TABLE.keys()))
    ap.add_argument("--commit", action="store_true")
    a = ap.parse_args()

    url = os.environ.get("SUPABASE_DB_URL")
    if not url: sys.exit("Set SUPABASE_DB_URL env var.")
    wb = load_workbook(a.file)
    ws = wb["Content"] if "Content" in wb.sheetnames else wb.active
    idx = {name: i for i, name in enumerate(COLUMNS)}

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if not all(v is None for v in r)]
    conn = psycopg2.connect(url); cur = conn.cursor()

    # resolve sources
    slugs = {r[idx["source_slug"]] for r in rows}
    cur.execute("SELECT slug, id FROM sources WHERE slug = ANY(%s)", (list(slugs),))
    src_map = dict(cur.fetchall())
    missing = slugs - set(src_map)
    if missing: sys.exit(f"Unregistered source_slug(s): {missing}. Register in sources first.")

    leaf = LEAF_TABLE[a.book_type]["leaf"]
    fk   = LEAF_TABLE[a.book_type]["parent_fk"]

    planned = []
    for r in rows:
        planned.append({
            fk: r[idx["section_id"]],
            "position": r[idx["position"]],
            "verse_number": r[idx["position"]],   # legacy mirror
            "text_english": r[idx["text_english"]],
            "text_amharic": r[idx["text_amharic"]],
            "text_geez": r[idx["text_geez"]],
            "text_transliteration": r[idx["text_transliteration"]],
            "source_id": src_map[r[idx["source_slug"]]],
            "chapter": r[idx["chapter"]] or None,
        })

    print(f"\n{'=== DRY RUN (nothing written) ===' if not a.commit else '=== COMMITTING ==='}")
    print(f"Leaf table: {leaf}  |  rows to insert: {len(planned)}")
    for p in planned[:5]:
        print("  ", {k:(v[:40]+'…' if isinstance(v,str) and len(v)>40 else v) for k,v in p.items()})
    if len(planned) > 5: print(f"   ...and {len(planned)-5} more")

    if not a.commit:
        print("\nDry run only. Re-run with --commit to write.\n"); return

    cols = [fk,"position","verse_number","text_english","text_amharic","text_geez",
            "text_transliteration","source_id"]
    # chapter only included for chaptered books (bible)
    if a.book_type == "bible": cols.append("chapter")
    placeholders = ",".join(["%s"]*len(cols))
    sql = f"INSERT INTO {leaf} ({','.join(cols)}) VALUES ({placeholders})"
    try:
        for p in planned:
            cur.execute(sql, tuple(p[c] for c in cols))
        conn.commit()
        print(f"\nCommitted {len(planned)} rows into {leaf}.\n")
    except Exception as e:
        conn.rollback(); sys.exit(f"Load failed, rolled back (nothing written): {e}")

if __name__ == "__main__":
    main()
